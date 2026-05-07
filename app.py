import os
import sys
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook, Workbook
from io import BytesIO
import secrets

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import config

app = Flask(__name__)
app.config.from_object(config['development'])  # Default to development

# Override with environment-specific config
if os.environ.get('FLASK_ENV') == 'production':
    app.config.from_object(config['production'])

# Ensure SECRET_KEY is set
if not app.config.get('SECRET_KEY') or app.config['SECRET_KEY'] == 'dev-secret-key-change-in-production':
    if os.environ.get('FLASK_ENV') == 'production':
        raise RuntimeError("SECRET_KEY must be set in production environment")
    app.config['SECRET_KEY'] = secrets.token_hex(32)

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL') or 'sqlite:///students.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# File paths
app.config['UPLOAD_FOLDER'] = 'data'
app.config['EXCEL_TEMPLATE'] = os.path.join(app.config['UPLOAD_FOLDER'], 'Template Current Students (1).xlsx')
app.config['EXCEL_OUTPUT'] = os.path.join(app.config['UPLOAD_FOLDER'], 'submissions.xlsx')

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'

# Ensure data directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ------------------ CSRF PROTECTION ------------------

def generate_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

def validate_csrf_token():
    token = session.get('_csrf_token')
    form_token = request.form.get('_csrf_token') or request.headers.get('X-CSRF-Token')
    if not token or not form_token or token != form_token:
        return False
    return True

def csrf_protect(f):
    """Decorator to validate CSRF token on POST requests."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method == 'POST' and not validate_csrf_token():
            # Return JSON for AJAX/API requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes['application/json']:
                return jsonify({'success': False, 'message': 'CSRF token invalid'}), 403
            flash('CSRF token invalid. Please try again.', 'error')
            return redirect(request.url or url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def before_request():
    # Generate CSRF token for all requests
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)

# Make csrf_token available to all templates
@app.context_processor
def inject_csrf():
    return dict(csrf_token=generate_csrf_token)

# ------------------ DATABASE MODELS ------------------

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), unique=True, nullable=False)  # dummy login
    password_hash = db.Column(db.String(256), nullable=False)
    full_name = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    submission = db.relationship('Submission', backref='student', lazy=True, uselist=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Submission(db.Model):
    __tablename__ = 'submission'
    __table_args__ = (db.UniqueConstraint('student_id', name='uq_submission_student_id'),)

    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    # Form fields matching Excel columns - ALL MANDATORY
    student_name = db.Column(db.String(200), nullable=False)
    father_name = db.Column(db.String(200), nullable=False)
    gender = db.Column(db.String(20), nullable=False)
    roll_no = db.Column(db.String(50), nullable=False)
    admission_date = db.Column(db.String(50), nullable=False)
    nationality = db.Column(db.String(100), nullable=False)
    cnic_number = db.Column(db.String(50), nullable=False)
    passport_number = db.Column(db.String(50), nullable=False)
    date_of_birth = db.Column(db.String(50), nullable=False)
    phone_number = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(200), nullable=False)
    domicile_district = db.Column(db.String(200), nullable=False)
    domicile_province = db.Column(db.String(200), nullable=False)
    mailing_address = db.Column(db.Text, nullable=False)
    city = db.Column(db.String(200), nullable=False)
    ssc_degree_name = db.Column(db.String(200), nullable=False)
    ssc_board_name = db.Column(db.String(200), nullable=False)
    ssc_total_marks = db.Column(db.String(50), nullable=False)
    ssc_obtained_marks = db.Column(db.String(50), nullable=False)
    hssc_degree_name = db.Column(db.String(200), nullable=False)
    hssc_degree_nomenclature = db.Column(db.String(50), nullable=False)
    hssc_board_name = db.Column(db.String(200), nullable=False)
    hssc_total_marks = db.Column(db.String(50), nullable=False)
    hssc_obtained_marks = db.Column(db.String(50), nullable=False)

class Admin(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(200), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

@login_manager.user_loader
def load_admin(admin_id):
    return Admin.query.get(int(admin_id))

# ------------------ EXCEL HANDLER ------------------

class ExcelHandler:
    EXCEL_COLUMNS = [
        'Student Name', 'FatherName', 'Gender', 'RollNo', 'Admission Date',
        'Nationality', 'CNIC Number', 'Passport Number', 'Date of Birth',
        'PhoneNumber', 'Email', 'Domicile District', 'Domicile Province',
        'Mailing Address', 'City (Mailing Address)', 'SSC Degree Name',
        'SSC Board Name', 'SSC Total Marks', 'SSC Obtained Marks',
        'HSSC Degree Name', 'HSSC Degree Nomenclature', 'HSSC Board Name',
        'HSSC Total Marks', 'HSSC Obtained Marks'
    ]

    @staticmethod
    def map_submission_to_row(submission):
        """Convert submission object to ordered list matching Excel columns."""
        return [
            submission.student_name,
            submission.father_name,
            submission.gender,
            submission.roll_no,
            submission.admission_date,
            submission.nationality,
            submission.cnic_number,
            submission.passport_number,
            submission.date_of_birth,
            submission.phone_number,
            submission.email,
            submission.domicile_district,
            submission.domicile_province,
            submission.mailing_address,
            submission.city,
            submission.ssc_degree_name,
            submission.ssc_board_name,
            submission.ssc_total_marks,
            submission.ssc_obtained_marks,
            submission.hssc_degree_name,
            submission.hssc_degree_nomenclature,
            submission.hssc_board_name,
            submission.hssc_total_marks,
            submission.hssc_obtained_marks,
        ]

    @staticmethod
    def load_template():
        """Load the Excel template, create if not exists."""
        if os.path.exists(app.config['EXCEL_TEMPLATE']):
            wb = load_workbook(app.config['EXCEL_TEMPLATE'])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.append(ExcelHandler.EXCEL_COLUMNS)
            wb.save(app.config['EXCEL_TEMPLATE'])
        return wb

    @staticmethod
    def save_submission(submission_data):
        """Append a new submission row to the Excel output file."""
        if os.path.exists(app.config['EXCEL_OUTPUT']):
            wb = load_workbook(app.config['EXCEL_OUTPUT'])
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            # Add headers first
            ws.append(ExcelHandler.EXCEL_COLUMNS)
        ws.append(submission_data)
        wb.save(app.config['EXCEL_OUTPUT'])
        return True

    @staticmethod
    def get_all_submissions():
        """Read all submission rows from output Excel."""
        if not os.path.exists(app.config['EXCEL_OUTPUT']):
            return []
        wb = load_workbook(app.config['EXCEL_OUTPUT'])
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            return []
        return rows[1:]  # Skip header

# ------------------ ROUTES ------------------

@app.route('/')
def index():
    """Landing page with option for student or admin login."""
    return render_template('index.html')

@app.route('/student/register', methods=['GET', 'POST'])
@csrf_protect
def student_register():
    """Generate dummy credentials for a new student."""
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        if not full_name:
            flash('Please provide a full name', 'error')
            return redirect(url_for('student_register'))

        # Generate unique username (email format) and random password
        base_username = full_name.lower().replace(' ', '.')
        username = f"{base_username}@bbsul.edu.pk"
        counter = 1
        while Student.query.filter_by(username=username).first():
            username = f"{base_username}{counter}@bbsul.edu.pk"
            counter += 1

        password = secrets.token_urlsafe(8)

        student = Student(username=username, full_name=full_name)
        student.set_password(password)
        db.session.add(student)
        db.session.commit()

        flash(f'Credentials generated for {full_name}', 'success')
        flash(f'Login: {username}', 'info')
        flash(f'Password: {password}', 'info')
        flash('Please write down these credentials. Student can now log in and fill the form once.', 'warning')
        return redirect(url_for('student_register'))

    students = Student.query.all()
    return render_template('register.html', students=students)

@app.route('/student/login', methods=['GET', 'POST'])
@csrf_protect
def student_login():
    """Student login to access the form."""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        student = Student.query.filter_by(username=username).first()
        if student and student.check_password(password):
            session['student_id'] = student.id
            session['student_name'] = student.full_name
            return redirect(url_for('student_form'))
        else:
            flash('Invalid credentials', 'error')

    return render_template('student_login.html')

@app.route('/student/form', methods=['GET', 'POST'])
@csrf_protect
def student_form():
    """Fillable form for students (one-time only)."""
    student_id = session.get('student_id')
    if not student_id:
        return redirect(url_for('student_login'))

    student = Student.query.get_or_404(student_id)

    # Check if already submitted
    if student.submission:
        flash('You have already submitted the form. Only one submission is allowed.', 'info')
        return render_template('already_submitted.html', student=student)

    if request.method == 'POST':
        # Validate ALL fields are now mandatory (24 fields)
        all_fields = [
            'student_name', 'father_name', 'gender', 'roll_no', 'admission_date',
            'nationality', 'cnic_number', 'passport_number', 'date_of_birth',
            'phone_number', 'email', 'domicile_district', 'domicile_province',
            'mailing_address', 'city', 'ssc_degree_name', 'ssc_board_name',
            'ssc_total_marks', 'ssc_obtained_marks', 'hssc_degree_name',
            'hssc_degree_nomenclature', 'hssc_board_name', 'hssc_total_marks',
            'hssc_obtained_marks'
        ]
        missing = [f for f in all_fields if not request.form.get(f, '').strip()]
        if missing:
            flash(f'All fields are required. Missing: {", ".join(missing)}', 'error')
            return redirect(url_for('student_form'))

        # Create submission
        submission = Submission(
            student_id=student.id,
            student_name=request.form.get('student_name'),
            father_name=request.form.get('father_name'),
            gender=request.form.get('gender'),
            roll_no=request.form.get('roll_no'),
            admission_date=request.form.get('admission_date'),
            nationality=request.form.get('nationality'),
            cnic_number=request.form.get('cnic_number'),
            passport_number=request.form.get('passport_number'),
            date_of_birth=request.form.get('date_of_birth'),
            phone_number=request.form.get('phone_number'),
            email=request.form.get('email'),
            domicile_district=request.form.get('domicile_district'),
            domicile_province=request.form.get('domicile_province'),
            mailing_address=request.form.get('mailing_address'),
            city=request.form.get('city'),
            ssc_degree_name=request.form.get('ssc_degree_name'),
            ssc_board_name=request.form.get('ssc_board_name'),
            ssc_total_marks=request.form.get('ssc_total_marks'),
            ssc_obtained_marks=request.form.get('ssc_obtained_marks'),
            hssc_degree_name=request.form.get('hssc_degree_name'),
            hssc_degree_nomenclature=request.form.get('hssc_degree_nomenclature'),
            hssc_board_name=request.form.get('hssc_board_name'),
            hssc_total_marks=request.form.get('hssc_total_marks'),
            hssc_obtained_marks=request.form.get('hssc_obtained_marks'),
        )
        db.session.add(submission)

        # Save to Excel
        excel_data = ExcelHandler.map_submission_to_row(submission)
        ExcelHandler.save_submission(excel_data)

        try:
            db.session.commit()
            flash('Form submitted successfully!', 'success')
            return redirect(url_for('student_form'))
        except IntegrityError:
            db.session.rollback()
            # Check if it's a duplicate student_id error
            existing = Submission.query.filter_by(student_id=student.id).first()
            if existing:
                flash('You have already submitted the form. Only one submission is allowed.', 'info')
                return render_template('already_submitted.html', student=student)
            else:
                flash('A database error occurred. Please try again.', 'error')
                return redirect(url_for('student_form'))

    return render_template('form.html', student=student)

@app.route('/student/logout')
def student_logout():
    session.clear()
    return redirect(url_for('index'))

# ------------------ ADMIN ROUTES ------------------

@app.route('/admin/login', methods=['GET', 'POST'])
@csrf_protect
def admin_login():
    """Admin login with predefined credentials."""
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()

        if email == 'rishabh@bbsul.edu.pk':
            admin = Admin.query.filter_by(email=email).first()
            if not admin:
                admin = Admin(email=email)
                admin.set_password('abc1234')
                db.session.add(admin)
                db.session.commit()

            if admin.check_password(password):
                login_user(admin)
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid password for admin account', 'error')
        else:
            flash('Invalid admin email', 'error')

    return render_template('admin_login.html')

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    """Admin view all submissions."""
    submissions = Submission.query.all()
    students = Student.query.all()
    return render_template('admin_dashboard.html', submissions=submissions, students=students)

@app.route('/admin/api/stats')
@login_required
def admin_stats_api():
    """Return real-time stats as JSON for AJAX updates."""
    total_submissions = db.session.query(db.func.count(Submission.id)).scalar() or 0
    total_students = db.session.query(db.func.count(Student.id)).scalar() or 0
    pending_submissions = total_students - total_submissions

    return jsonify({
        'total_submissions': total_submissions,
        'total_students': total_students,
        'pending_submissions': pending_submissions
    })

@app.route('/admin/download-excel')
@login_required
def download_excel():
    """Download the Excel file with all submissions."""
    if os.path.exists(app.config['EXCEL_OUTPUT']):
        return send_file(app.config['EXCEL_OUTPUT'], as_attachment=True, download_name='student_submissions.xlsx')
    else:
        flash('No submissions found. Excel file not created yet.', 'warning')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/admin/delete-submission/<int:submission_id>', methods=['POST'])
@login_required
@csrf_protect
def delete_submission(submission_id):
    """Delete a single submission via AJAX."""
    submission = Submission.query.get_or_404(submission_id)
    db.session.delete(submission)

    try:
        # Rebuild Excel first (reads current DB state, excluding this deletion)
        rebuild_excel()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

    # Return current stats
    total_submissions = db.session.query(db.func.count(Submission.id)).scalar() or 0
    total_students = db.session.query(db.func.count(Student.id)).scalar() or 0
    pending_submissions = total_students - total_submissions

    return jsonify({
        'success': True,
        'message': 'Submission deleted',
        'stats': {
            'total_submissions': total_submissions,
            'total_students': total_students,
            'pending_submissions': pending_submissions
        }
    })

@app.route('/admin/flush-all', methods=['POST'])
@login_required
@csrf_protect
def flush_all_data():
    """Delete all submissions and student data, then reset Excel file."""
    from sqlalchemy import text

    # Delete all submissions first (to satisfy foreign key constraints)
    Submission.query.delete()

    # Delete all students
    Student.query.delete()

    try:
        # Reset Excel file with only headers
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(ExcelHandler.EXCEL_COLUMNS)
        wb.save(app.config['EXCEL_OUTPUT'])
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

    # Return updated stats
    total_submissions = db.session.query(db.func.count(Submission.id)).scalar() or 0
    total_students = db.session.query(db.func.count(Student.id)).scalar() or 0
    pending_submissions = total_students - total_submissions

    return jsonify({
        'success': True,
        'message': 'All student data and submissions cleared',
        'stats': {
            'total_submissions': total_submissions,
            'total_students': total_students,
            'pending_submissions': pending_submissions
        }
    })

def rebuild_excel():
    """Rebuild the Excel file from all current submissions in database."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(ExcelHandler.EXCEL_COLUMNS)

    submissions = Submission.query.order_by(Submission.id).all()
    for sub in submissions:
        row = ExcelHandler.map_submission_to_row(sub)
        ws.append(row)

    wb.save(app.config['EXCEL_OUTPUT'])

# ------------------ INITIALIZATION ------------------

def init_db():
    with app.app_context():
        # Check if migration needed (unique constraint or nullable fields)
        try:
            inspector = db.inspect(db.engine)
            columns = inspector.get_columns('submission')
            constraints = inspector.get_unique_constraints('submission')

            # Check unique constraint
            has_unique = any('student_id' in str(c['column_names']) for c in constraints)

            # Check if any nullable columns exist (excluding id)
            has_nullable = any(c['nullable'] for c in columns if c['name'] not in ('id', 'student_id', 'submitted_at'))

            needs_migration = not has_unique or has_nullable

            if needs_migration:
                print("Migrating database: enforcing unique constraint and NOT NULL on all fields...")
                migrate_submission_table()
            else:
                print("Database schema is up to date (all fields mandatory).")
        except Exception as e:
            # Table might not exist yet
            print(f"Schema check skipped: {e}")

        db.create_all()
        # Ensure output Excel file exists with header
        if not os.path.exists(app.config['EXCEL_OUTPUT']):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.append(ExcelHandler.EXCEL_COLUMNS)
            wb.save(app.config['EXCEL_OUTPUT'])

def migrate_submission_table():
    """Migrate submission table to add UNIQUE constraint and NOT NULL on all fields."""
    from sqlalchemy import text

    # Fetch all existing submissions
    submissions = Submission.query.all()
    print(f"Migrating {len(submissions)} submissions...")

    # Helper to safely convert None to empty string
    def safe(value):
        return '' if value is None else str(value)

    # Store data with defaults for NULL
    old_data = []
    null_counts = {}
    for sub in submissions:
        row = {
            'id': sub.id,
            'student_id': sub.student_id,
            'submitted_at': sub.submitted_at,
            'student_name': safe(sub.student_name),
            'father_name': safe(sub.father_name),
            'gender': safe(sub.gender),
            'roll_no': safe(sub.roll_no),
            'admission_date': safe(sub.admission_date),
            'nationality': safe(sub.nationality),
            'cnic_number': safe(sub.cnic_number),
            'passport_number': safe(sub.passport_number),
            'date_of_birth': safe(sub.date_of_birth),
            'phone_number': safe(sub.phone_number),
            'email': safe(sub.email),
            'domicile_district': safe(sub.domicile_district),
            'domicile_province': safe(sub.domicile_province),
            'mailing_address': safe(sub.mailing_address),
            'city': safe(sub.city),
            'ssc_degree_name': safe(sub.ssc_degree_name),
            'ssc_board_name': safe(sub.ssc_board_name),
            'ssc_total_marks': safe(sub.ssc_total_marks),
            'ssc_obtained_marks': safe(sub.ssc_obtained_marks),
            'hssc_degree_name': safe(sub.hssc_degree_name),
            'hssc_degree_nomenclature': safe(sub.hssc_degree_nomenclature),
            'hssc_board_name': safe(sub.hssc_board_name),
            'hssc_total_marks': safe(sub.hssc_total_marks),
            'hssc_obtained_marks': safe(sub.hssc_obtained_marks),
        }
        old_data.append(row)

    # Recreate table with all NOT NULL constraints
    with db.engine.connect() as conn:
        conn.execute(text("DROP TABLE IF EXISTS submission_new"))
        conn.execute(text("""
            CREATE TABLE submission_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL UNIQUE,
                submitted_at DATETIME,
                student_name VARCHAR(200) NOT NULL,
                father_name VARCHAR(200) NOT NULL,
                gender VARCHAR(20) NOT NULL,
                roll_no VARCHAR(50) NOT NULL,
                admission_date VARCHAR(50) NOT NULL,
                nationality VARCHAR(100) NOT NULL,
                cnic_number VARCHAR(50) NOT NULL,
                passport_number VARCHAR(50) NOT NULL,
                date_of_birth VARCHAR(50) NOT NULL,
                phone_number VARCHAR(50) NOT NULL,
                email VARCHAR(200) NOT NULL,
                domicile_district VARCHAR(200) NOT NULL,
                domicile_province VARCHAR(200) NOT NULL,
                mailing_address TEXT NOT NULL,
                city VARCHAR(200) NOT NULL,
                ssc_degree_name VARCHAR(200) NOT NULL,
                ssc_board_name VARCHAR(200) NOT NULL,
                ssc_total_marks VARCHAR(50) NOT NULL,
                ssc_obtained_marks VARCHAR(50) NOT NULL,
                hssc_degree_name VARCHAR(200) NOT NULL,
                hssc_degree_nomenclature VARCHAR(50) NOT NULL,
                hssc_board_name VARCHAR(200) NOT NULL,
                hssc_total_marks VARCHAR(50) NOT NULL,
                hssc_obtained_marks VARCHAR(50) NOT NULL,
                FOREIGN KEY(student_id) REFERENCES student (id)
            )
        """))
        conn.commit()

    # Re-insert data, keep first per student only
    seen = set()
    duplicate_count = 0
    for data in old_data:
        if data['student_id'] in seen:
            duplicate_count += 1
            continue
        stmt = text("""
            INSERT INTO submission_new (
                id, student_id, submitted_at, student_name, father_name, gender, roll_no,
                admission_date, nationality, cnic_number, passport_number, date_of_birth,
                phone_number, email, domicile_district, domicile_province, mailing_address,
                city, ssc_degree_name, ssc_board_name, ssc_total_marks, ssc_obtained_marks,
                hssc_degree_name, hssc_degree_nomenclature, hssc_board_name,
                hssc_total_marks, hssc_obtained_marks
            ) VALUES (
                :id, :student_id, :submitted_at, :student_name, :father_name, :gender,
                :roll_no, :admission_date, :nationality, :cnic_number, :passport_number,
                :date_of_birth, :phone_number, :email, :domicile_district, :domicile_province,
                :mailing_address, :city, :ssc_degree_name, :ssc_board_name, :ssc_total_marks,
                :ssc_obtained_marks, :hssc_degree_name, :hssc_degree_nomenclature,
                :hssc_board_name, :hssc_total_marks, :hssc_obtained_marks
            )
        """)
        with db.engine.connect() as conn:
            conn.execute(stmt, data)
            conn.commit()
        seen.add(data['student_id'])

    # Replace table
    with db.engine.connect() as conn:
        conn.execute(text("DROP TABLE submission"))
        conn.execute(text("ALTER TABLE submission_new RENAME TO submission"))
        conn.commit()

    print(f"Migration complete: {len(seen)} submissions migrated.")
    if duplicate_count > 0:
        print(f"  Removed {duplicate_count} duplicates.")
    print("  All fields now mandatory (NOT NULL).")
    print("  UNIQUE constraint on student_id enforced.")

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
